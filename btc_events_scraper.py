"""
Bitcoin Price Events Scraper
Collects major historical events that influenced Bitcoin's price
and stores them in an Excel file for price prediction analysis.
"""

import time
import re
import logging
from datetime import datetime
from dataclasses import dataclass, field

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

OUTPUT_FILE = "bitcoin_events.xlsx"


@dataclass
class BTCEvent:
    date: str
    title: str
    description: str
    category: str
    impact: str          # BULLISH / BEARISH / NEUTRAL
    source: str
    price_before: str = ""
    price_after: str = ""
    price_change_pct: str = ""
    url: str = ""
    tags: str = ""


# ---------------------------------------------------------------------------
# CURATED HISTORICAL DATASET
# A high-quality seed of every major documented event + approximate prices.
# Sources: CoinDesk, Investopedia, Wikipedia, Bitcoin Magazine archives.
# ---------------------------------------------------------------------------

HISTORICAL_EVENTS: list[BTCEvent] = [
    # 2008
    BTCEvent("2008-10-31", "Bitcoin Whitepaper Published",
             "Satoshi Nakamoto publishes 'Bitcoin: A Peer-to-Peer Electronic Cash System' "
             "on the cryptography mailing list, introducing decentralized digital currency.",
             "Technology", "BULLISH", "Bitcoin Whitepaper",
             "0", "0", "0", "", "satoshi,whitepaper,creation,origin"),
    BTCEvent("2008-11-09", "Bitcoin Source Code Released to Mailing List",
             "Satoshi shares the full Bitcoin source code with the cypherpunk mailing list "
             "for peer review.",
             "Technology", "BULLISH", "Bitcoin Mailing List",
             "0", "0", "0", "", "satoshi,source code,cypherpunk"),

    # 2009
    BTCEvent("2009-01-03", "Bitcoin Genesis Block Mined",
             "Satoshi Nakamoto mines the first Bitcoin block (Block 0), embedding "
             "'The Times 03/Jan/2009 Chancellor on brink of second bailout for banks'. "
             "50 BTC reward is unspendable by convention.",
             "Technology", "BULLISH", "Bitcoin Whitepaper / Genesis Block",
             "0", "0", "0", "", "genesis,satoshi,creation,block 0"),
    BTCEvent("2009-01-09", "Bitcoin v0.1 Released — Network Goes Live",
             "Satoshi Nakamoto releases Bitcoin v0.1 on SourceForge. The Bitcoin network "
             "officially goes live with the first nodes mining.",
             "Technology", "BULLISH", "Bitcoin Talk / SourceForge",
             "0", "0", "0", "", "v0.1,software release,network launch"),
    BTCEvent("2009-01-12", "First Bitcoin Transaction",
             "Satoshi sends 10 BTC to Hal Finney in the first ever peer-to-peer "
             "Bitcoin transaction (Block 170).",
             "Adoption", "BULLISH", "Bitcoin History",
             "~0", "~0", "0", "", "first transaction,hal finney,block 170"),
    BTCEvent("2009-10-05", "First Bitcoin Exchange Rate Established",
             "New Liberty Standard publishes the first Bitcoin exchange rate: $1 = 1,309 BTC, "
             "calculated from electricity cost to mine Bitcoin.",
             "Exchange", "BULLISH", "New Liberty Standard",
             "0.00076", "0.00076", "0", "", "exchange rate,first price,new liberty standard"),
    BTCEvent("2009-10-12", "First BTC/USD Trade Recorded",
             "Martti Malmi sells 5,050 BTC for $5.02 via PayPal — the first documented "
             "USD sale of Bitcoin.",
             "Exchange", "BULLISH", "Bitcoin Talk / NLS",
             "0.00099", "0.00099", "0", "", "first sale,usd,paypal,martti malmi"),

    # 2010
    BTCEvent("2010-02-06", "Bitcoin Market (First Exchange) Launches",
             "The Bitcoin Market website launches as the first dedicated exchange platform, "
             "allowing users to trade BTC directly for USD via PayPal.",
             "Exchange", "BULLISH", "Bitcoin Talk",
             "0.003", "0.004", "+33", "", "exchange,bitcoin market,first exchange"),
    BTCEvent("2010-05-22", "Bitcoin Pizza Day",
             "Laszlo Hanyecz pays 10,000 BTC for two pizzas — the first real-world "
             "commercial transaction. Price ~$0.0025/BTC. "
             "Those coins would be worth ~$1B at 2024 prices.",
             "Adoption", "BULLISH", "Bitcoin Talk Forum",
             "~0.0025", "~0.0025", "0", "", "pizza,first purchase,adoption,laszlo"),
    BTCEvent("2010-07-11", "Slashdot Article Triggers First Price Spike",
             "Slashdot publishes 'Bitcoin Releases Version 0.3'. Price rises 10x from "
             "$0.008 to $0.08 within days as wave of new users arrives.",
             "Media", "BULLISH", "Slashdot",
             "0.008", "0.08", "+900", "", "slashdot,media,spike,adoption"),
    BTCEvent("2010-07-17", "Mt. Gox Exchange Launches",
             "Jed McCaleb launches Mt. Gox as a Bitcoin exchange. Dramatically increases "
             "liquidity and accessibility. Price jumps from $0.05 to $0.08 in days.",
             "Exchange", "BULLISH", "Mt. Gox / CoinDesk",
             "0.05", "0.08", "+60", "", "exchange,mt gox,liquidity,jed mccaleb"),
    BTCEvent("2010-08-15", "Bitcoin Value Overflow Incident",
             "A bug creates 184 billion BTC in Block 74638. Satoshi patches the bug; "
             "the chain is hard-forked within hours. First serious protocol bug.",
             "Security", "BEARISH", "Bitcoin Talk / CVE",
             "0.07", "0.05", "-28", "", "bug,overflow,hard fork,satoshi"),
    BTCEvent("2010-11-06", "Bitcoin Breaks $0.50",
             "Bitcoin price crosses $0.50 for the first time as adoption and exchange "
             "infrastructure grows significantly.",
             "Price Milestone", "BULLISH", "Mt. Gox",
             "0.35", "0.50", "+43", "", "milestone,$0.50,price"),

    # 2011
    BTCEvent("2011-02-09", "Bitcoin Reaches $1",
             "Bitcoin hits $1 for the first time on Mt. Gox — parity with the US dollar. "
             "Widely covered by tech media.",
             "Price Milestone", "BULLISH", "Mt. Gox / TIME Magazine",
             "0.30", "1.00", "+233", "", "milestone,$1,dollar parity"),
    BTCEvent("2011-04-16", "TIME Magazine Article on Bitcoin",
             "TIME Magazine publishes 'Online Cash Bitcoin Could Challenge Governments, "
             "Banks', massively increasing public awareness. Price surges from $1 to $8.",
             "Media", "BULLISH", "TIME Magazine",
             "0.90", "8.00", "+789", "", "media,mainstream,time"),
    BTCEvent("2011-06-02", "Bitcoin Reaches All-Time High of $31.91",
             "Bitcoin surges to $31.91 on Mt. Gox — its highest price to date — driven by "
             "a Gawker article about Silk Road and surging mainstream interest.",
             "Price Milestone", "BULLISH", "Mt. Gox / Gawker",
             "8.00", "31.91", "+299", "", "all time high,ath,$31,gawker,silk road"),
    BTCEvent("2011-06-13", "Gawker Article on Silk Road",
             "Gawker publishes 'The Underground Website Where You Can Buy Any Drug', "
             "exposing Silk Road and Bitcoin to millions. Price spikes then draws scrutiny.",
             "Media / Regulation", "NEUTRAL", "Gawker",
             "25.00", "31.91", "+28", "", "silk road,gawker,drugs,darknet,media"),
    BTCEvent("2011-06-19", "Mt. Gox Hack — Price Crashes from $17 to $0.01",
             "Mt. Gox is hacked; attacker manipulates the order book, crashing price from "
             "$17.50 to $0.01 momentarily. 60,000 user records stolen.",
             "Security / Hack", "BEARISH", "Mt. Gox / Bitcoin Talk",
             "17.50", "7.00", "-60", "", "hack,mt gox,security,data breach"),
    BTCEvent("2011-09-27", "Bitcoin Bottoms at $2.04 — First Major Bear Market",
             "Bitcoin crashes from $31.91 ATH to $2.04 (−94%) over the summer, "
             "its first extended bear market after the initial speculative bubble.",
             "Market", "BEARISH", "Mt. Gox",
             "6.00", "2.04", "-66", "", "bear market,bottom,crash,bubble"),
    BTCEvent("2011-11-14", "Litecoin Launched",
             "Charlie Lee launches Litecoin as the 'silver to Bitcoin's gold', "
             "the first major altcoin. Uses Scrypt PoW algorithm.",
             "Competition", "NEUTRAL", "Litecoin / Bitcoin Talk",
             "2.50", "2.40", "-4", "", "litecoin,altcoin,competition,charlie lee"),
    BTCEvent("2011-11-20", "Silk Road Attention from US Senators",
             "Senators Schumer and Manchin call on the DEA to shut down Silk Road, the "
             "first high-profile regulatory threat involving Bitcoin.",
             "Regulation", "BEARISH", "US Senate / CNN",
             "3.00", "2.20", "-26", "", "silk road,regulation,darknet,senate"),

    # 2012
    BTCEvent("2012-03-01", "Linode Hack — 46,000 BTC Stolen",
             "Hackers compromise Linode and steal 46,703 BTC from several Bitcoin "
             "businesses including Bitcoinica and Slush Pool.",
             "Security / Hack", "BEARISH", "Linode / Bitcoin Talk",
             "5.00", "4.50", "-10", "", "linode,hack,security,bitcoinica"),
    BTCEvent("2012-05-11", "Bitcoinica Hack — 18,000 BTC Stolen",
             "Bitcoinica, a major Bitcoin trading platform, is hacked for 18,548 BTC. "
             "Platform shuts down entirely after a second hack.",
             "Security / Hack", "BEARISH", "Bitcoinica / Bitcoin Talk",
             "5.50", "5.00", "-9", "", "bitcoinica,hack,exchange,security"),
    BTCEvent("2012-08-15", "Bitcoin Foundation Established",
             "The Bitcoin Foundation is founded to standardize, protect, and promote "
             "Bitcoin. First major governance/advocacy organization for the protocol.",
             "Adoption / Governance", "BULLISH", "Bitcoin Foundation",
             "10.00", "11.00", "+10", "", "bitcoin foundation,governance,advocacy"),
    BTCEvent("2012-09-27", "Coinbase Founded",
             "Brian Armstrong and Fred Ehrsam found Coinbase, which will become the largest "
             "US cryptocurrency exchange.",
             "Exchange / Adoption", "BULLISH", "Coinbase",
             "11.50", "12.00", "+4", "", "coinbase,exchange,adoption,founded"),
    BTCEvent("2012-11-28", "First Bitcoin Halving",
             "Block reward halves from 50 BTC to 25 BTC at block 210,000. Historically "
             "the start of the first major bull cycle.",
             "Halving", "BULLISH", "Bitcoin Protocol",
             "12.00", "13.00", "+8", "", "halving,supply,mining,block 210000"),
    BTCEvent("2012-12-27", "Wordpress Accepts Bitcoin",
             "Wordpress becomes one of the first major websites to accept Bitcoin as "
             "payment — a significant mainstream adoption milestone.",
             "Adoption", "BULLISH", "Wordpress / Bitcoin Magazine",
             "13.00", "13.50", "+4", "", "wordpress,adoption,payment,merchant"),

    # 2013
    BTCEvent("2013-02-19", "Reddit Accepts Bitcoin for Gold",
             "Reddit announces it accepts Bitcoin for its Reddit Gold subscription, "
             "one of the first major social platforms to accept crypto.",
             "Adoption", "BULLISH", "Reddit",
             "27.00", "28.00", "+4", "", "reddit,adoption,payment,social media"),
    BTCEvent("2013-03-18", "Cyprus Bailout / Banking Crisis",
             "EU imposes a 10% levy on Cyprus bank deposits. Bitcoin seen as a safe haven "
             "and surges from $45 to $266 in weeks. First macro-driven rally.",
             "Macro / Banking", "BULLISH", "Reuters / CoinDesk",
             "45", "266", "+491", "", "banking crisis,safe haven,cyprus,macro"),
    BTCEvent("2013-03-25", "FinCEN Issues First Bitcoin Guidance",
             "US FinCEN classifies Bitcoin exchangers and administrators as money services "
             "businesses (MSBs) — first US regulatory framework for crypto.",
             "Regulation", "NEUTRAL", "FinCEN / US Treasury",
             "75", "72", "-4", "", "fincen,regulation,msb,compliance,aml"),
    BTCEvent("2013-04-10", "First Major Crash from $266",
             "After a parabolic rise Bitcoin crashes 83% in 2 days due to Mt. Gox "
             "overload, DDoS attacks, and profit taking.",
             "Market", "BEARISH", "Mt. Gox / CoinDesk",
             "266", "54", "-80", "", "crash,bubble,mt gox,ddos"),
    BTCEvent("2013-05-07", "Coinbase Opens to General Public",
             "Coinbase removes its waitlist and opens to all US users, dramatically "
             "lowering the barrier to Bitcoin entry.",
             "Exchange / Adoption", "BULLISH", "Coinbase",
             "105", "115", "+10", "", "coinbase,adoption,exchange,public"),
    BTCEvent("2013-05-28", "DHS Seizes Mt. Gox US Bank Account",
             "Department of Homeland Security seizes funds from Mt. Gox's US Dwolla account "
             "for failure to register as an MSB. First US enforcement action.",
             "Regulation / Exchange", "BEARISH", "DHS / Reuters",
             "130", "110", "-15", "", "dhs,mt gox,seizure,regulation,enforcement"),
    BTCEvent("2013-08-06", "US Senate Holds First Bitcoin Hearing",
             "US Senate holds its first hearing on Bitcoin: 'Beyond Silk Road: Potential "
             "Risks, Threats, and Promises of Virtual Currencies'.",
             "Regulation", "NEUTRAL", "US Senate",
             "105", "120", "+14", "", "senate hearing,regulation,us government"),
    BTCEvent("2013-10-01", "Silk Road Shut Down by FBI",
             "FBI seizes Silk Road and arrests Ross Ulbricht. 26,000 BTC confiscated. "
             "Short-term bearish but long-term neutral as it removes criminal stigma.",
             "Regulation / Law", "BEARISH", "FBI / Reuters",
             "140", "109", "-22", "", "silk road,fbi,regulation,ross ulbricht"),
    BTCEvent("2013-10-14", "Baidu Accepts Bitcoin",
             "Chinese internet giant Baidu announces it will accept Bitcoin for its "
             "Jiasule web security service — signaling massive Chinese adoption.",
             "Adoption", "BULLISH", "Baidu / Reuters",
             "130", "145", "+12", "", "baidu,china,adoption,payment"),
    BTCEvent("2013-10-31", "World's First Bitcoin ATM Opens in Vancouver",
             "Robocoin installs the world's first Bitcoin ATM in a Vancouver coffee shop, "
             "marking a milestone in physical Bitcoin infrastructure.",
             "Adoption", "BULLISH", "Robocoin / Bitcoin Magazine",
             "195", "200", "+3", "", "bitcoin atm,robocoin,infrastructure,adoption"),
    BTCEvent("2013-11-18", "US Senate Hearing Positive on Bitcoin",
             "US Senate Commerce Committee hears positive testimony from federal officials "
             "calling Bitcoin a 'legitimate financial service'. Price surges $500 to $900.",
             "Regulation", "BULLISH", "US Senate / CoinDesk",
             "500", "900", "+80", "", "senate,regulation,positive,legitimate"),
    BTCEvent("2013-11-29", "Bitcoin Reaches $1,000",
             "Bitcoin hits $1,000 for the first time on Mt. Gox. BTC-China briefly "
             "surpasses Mt. Gox in volume. Widespread mainstream media coverage.",
             "Price Milestone", "BULLISH", "CoinDesk / CNN",
             "200", "1242", "+521", "", "milestone,$1000,all time high,china,btcchina"),
    BTCEvent("2013-12-05", "China PBOC Bans Banks from Bitcoin",
             "People's Bank of China prohibits financial institutions from handling Bitcoin "
             "transactions, triggering a crash from $1,200 to $600.",
             "Regulation", "BEARISH", "PBOC / Bloomberg",
             "1100", "600", "-45", "", "china,regulation,ban,pboc"),
    BTCEvent("2013-12-18", "Alibaba / Taobao Ban Bitcoin Transactions",
             "Alibaba's Taobao marketplace bans sale of Bitcoin and related products "
             "following the PBOC directive.",
             "Regulation / Adoption", "BEARISH", "Alibaba / Reuters",
             "650", "580", "-11", "", "alibaba,taobao,china,ban,regulation"),

    # 2014
    BTCEvent("2014-01-09", "Overstock.com Becomes First Major Retailer to Accept Bitcoin",
             "Overstock.com begins accepting Bitcoin — the first major US online retailer "
             "to do so. Sales hit $1M in the first month.",
             "Adoption", "BULLISH", "Overstock / CoinDesk",
             "950", "990", "+4", "", "overstock,adoption,retailer,ecommerce"),
    BTCEvent("2014-02-07", "Mt. Gox Halts Withdrawals",
             "Mt. Gox suspends all Bitcoin withdrawals citing 'transaction malleability'. "
             "Marks the beginning of its collapse. ~$500M in customer funds at risk.",
             "Exchange / Hack", "BEARISH", "Mt. Gox",
             "800", "550", "-31", "", "mt gox,withdrawal,collapse,transaction malleability"),
    BTCEvent("2014-02-24", "Mt. Gox Files Bankruptcy",
             "Mt. Gox files for bankruptcy after losing ~850,000 BTC (~$450M). "
             "Largest Bitcoin exchange collapse to date.",
             "Exchange / Hack", "BEARISH", "Mt. Gox / WSJ",
             "550", "420", "-23", "", "mt gox,bankruptcy,hack,850000 btc"),
    BTCEvent("2014-03-25", "IRS Classifies Bitcoin as Property",
             "US IRS rules Bitcoin is treated as property (not currency) for tax purposes, "
             "creating capital gains tax obligations for all transactions.",
             "Regulation / Tax", "BEARISH", "IRS / US Treasury",
             "560", "490", "-12", "", "irs,tax,property,capital gains,regulation"),
    BTCEvent("2014-06-27", "US Marshals Auction 29,656 BTC from Silk Road",
             "US Marshals Service auctions Silk Road-seized BTC. Tim Draper wins and "
             "announces he will support Bitcoin businesses.",
             "Regulation / Government", "BULLISH", "US Marshals / Reuters",
             "580", "610", "+5", "", "us marshals,auction,silk road,tim draper,government"),
    BTCEvent("2014-07-18", "Dell Accepts Bitcoin",
             "Dell becomes the largest retailer to accept Bitcoin, alongside Newegg — "
             "signaling mainstream merchant adoption.",
             "Adoption", "BULLISH", "Dell / Bitcoin Magazine",
             "615", "635", "+3", "", "dell,adoption,retailer,mainstream,ecommerce"),
    BTCEvent("2014-10-06", "Bearish Trend Deepens — Below $300",
             "Bitcoin falls below $300 as the prolonged post-Mt. Gox bear market continues. "
             "Mining becomes unprofitable for many operators.",
             "Market", "BEARISH", "CoinDesk",
             "420", "290", "-31", "", "bear market,mt gox,regulation,mining"),
    BTCEvent("2014-12-01", "Microsoft Accepts Bitcoin",
             "Microsoft begins accepting Bitcoin for digital content on Xbox and Windows "
             "Phone — one of the biggest mainstream adoption milestones.",
             "Adoption", "BULLISH", "Microsoft / CoinDesk",
             "370", "380", "+3", "", "microsoft,adoption,payment,xbox,mainstream"),

    # 2015
    BTCEvent("2015-01-05", "Bitstamp Hacked — 19,000 BTC Stolen",
             "Bitstamp exchange is hacked for 19,000 BTC (~$5.1M) via employee phishing. "
             "Briefly suspends trading.",
             "Security / Hack", "BEARISH", "Bitstamp / Reuters",
             "275", "255", "-7", "", "bitstamp,hack,security,phishing"),
    BTCEvent("2015-01-14", "Coinbase Raises $75M Series C",
             "Coinbase raises $75M from NYSE, BBVA, and others — institutional money "
             "enters Bitcoin at scale. Valuation reaches $400M.",
             "Institutional", "BULLISH", "Coinbase / Forbes",
             "175", "215", "+23", "", "coinbase,institutional,funding,series c"),
    BTCEvent("2015-01-26", "Bitcoin Bottoms at $152",
             "Bitcoin hits a multi-year low of $152.40, an 88% decline from the 2013 "
             "all-time high of $1,242.",
             "Market", "BEARISH", "Bitstamp",
             "200", "152", "-24", "", "bottom,bear market,low,2015"),
    BTCEvent("2015-05-01", "BitLicense Finalized in New York",
             "New York DFS finalizes the BitLicense framework — the first comprehensive "
             "US state-level crypto regulation. Many companies exit New York.",
             "Regulation", "BEARISH", "NYDFS / Bloomberg",
             "230", "210", "-9", "", "bitlicense,new york,nydfs,regulation,compliance"),
    BTCEvent("2015-07-30", "Ethereum Mainnet Launches",
             "Ethereum launches its Frontier mainnet, introducing smart contracts. "
             "Creates competition narrative but expands the overall crypto market.",
             "Competition / Technology", "NEUTRAL", "Ethereum Foundation",
             "280", "285", "+2", "", "ethereum,smart contracts,mainnet,frontier,competition"),
    BTCEvent("2015-08-01", "Bitcoin XT Fork Proposal",
             "Bitcoin XT proposes increasing block size to 8MB, creating deep community "
             "discord and scaling debate.",
             "Technology / Fork", "BEARISH", "Bitcoin Magazine",
             "280", "260", "-7", "", "fork,block size,scaling,bitcoin xt,gavin andresen"),
    BTCEvent("2015-10-22", "EU Court Rules Bitcoin Exempt from VAT",
             "European Court of Justice rules Bitcoin transactions are exempt from VAT, "
             "providing important regulatory clarity in Europe.",
             "Regulation", "BULLISH", "EU Court / Reuters",
             "265", "280", "+6", "", "eu,europe,vat,regulation,tax,court"),
    BTCEvent("2015-10-31", "The Economist: 'The Trust Machine'",
             "The Economist publishes a landmark cover story on blockchain, greatly "
             "increasing mainstream institutional awareness of Bitcoin.",
             "Media", "BULLISH", "The Economist",
             "320", "335", "+5", "", "economist,media,mainstream,blockchain,institutional"),

    # 2016
    BTCEvent("2016-01-14", "Mike Hearn Declares Bitcoin a Failed Experiment",
             "Core developer Mike Hearn writes a widely-read blog post declaring Bitcoin "
             "has 'failed' due to block size debate. Market drops ~20%.",
             "Technology / Media", "BEARISH", "Medium / Mike Hearn",
             "460", "370", "-20", "", "mike hearn,block size,scaling,community,failed"),
    BTCEvent("2016-05-02", "Craig Wright Claims to be Satoshi",
             "Australian entrepreneur Craig Wright publicly claims to be Bitcoin creator "
             "Satoshi Nakamoto but fails to provide cryptographic proof.",
             "Media / Technology", "NEUTRAL", "BBC / The Economist",
             "455", "450", "-1", "", "craig wright,satoshi,identity,controversy"),
    BTCEvent("2016-05-22", "Ethereum DAO Raises $150M",
             "The DAO raises $150M in Ether, diverting attention and capital from Bitcoin "
             "and demonstrating smart contract fundraising at scale.",
             "Competition", "NEUTRAL", "CoinDesk / Ethereum",
             "450", "440", "-2", "", "ethereum,dao,competition,smart contracts"),
    BTCEvent("2016-06-17", "Ethereum DAO Hack — $60M Stolen",
             "The DAO is hacked via reentrancy vulnerability; $60M in ETH drained. "
             "Ethereum hard forks to reverse — creating ETH/ETC split. Bitcoin benefits.",
             "Security / Competition", "BULLISH", "CoinDesk / Ethereum",
             "710", "730", "+3", "", "ethereum,dao hack,reentrancy,hard fork,etc"),
    BTCEvent("2016-07-09", "Second Bitcoin Halving",
             "Block reward halves from 25 BTC to 12.5 BTC at block 420,000. Bull cycle "
             "begins; price rises from ~$650 to $20,000 over the next 18 months.",
             "Halving", "BULLISH", "Bitcoin Protocol",
             "650", "660", "+2", "", "halving,supply,mining,block 420000"),
    BTCEvent("2016-08-02", "Bitfinex Hack — 119,756 BTC Stolen",
             "Bitfinex exchange is hacked for 119,756 BTC (~$72M) via multisig exploit. "
             "Price drops 20%. BFX tokens issued to affected users.",
             "Security / Hack", "BEARISH", "Bitfinex / Reuters",
             "600", "480", "-20", "", "hack,bitfinex,security,multisig"),
    BTCEvent("2016-11-12", "Bitcoin Breaks $700 — Post-Election Rally",
             "Bitcoin rallies past $700 as global uncertainty following the US election "
             "drives demand for non-correlated assets.",
             "Macro / Political", "BULLISH", "CoinDesk",
             "680", "710", "+4", "", "election,political,safe haven,macro"),

    # 2017
    BTCEvent("2017-01-05", "China PBOC Visits Exchanges — Leverage Crackdown",
             "PBOC visits major Chinese exchanges (OKCoin, Huobi, BTCC) over leverage "
             "trading concerns. Temporary withdrawal freezes imposed.",
             "Regulation", "BEARISH", "PBOC / Bloomberg",
             "1130", "900", "-20", "", "china,pboc,exchange,regulation,leverage"),
    BTCEvent("2017-03-10", "SEC Rejects Winklevoss Bitcoin ETF",
             "SEC rejects the first Bitcoin ETF application by the Winklevoss twins, "
             "citing lack of market regulation.",
             "Regulation / ETF", "BEARISH", "SEC / Bloomberg",
             "1290", "1050", "-19", "", "etf,sec,regulation,winklevoss"),
    BTCEvent("2017-03-25", "Bitcoin Reclaims $1,000 After 4 Years",
             "Bitcoin breaks $1,000 for the first time since the 2013 bubble, "
             "marking the true start of the 2017 bull run.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "950", "1050", "+10", "", "milestone,$1000,bull run,recovery"),
    BTCEvent("2017-05-25", "Japan Recognizes Bitcoin as Legal Tender",
             "Japan's Payment Services Act officially recognizes Bitcoin as a legal payment "
             "method, triggering a major rally.",
             "Regulation / Adoption", "BULLISH", "Japanese FSA / Reuters",
             "2200", "2700", "+23", "", "japan,legal tender,adoption,regulation"),
    BTCEvent("2017-08-01", "Bitcoin Cash Hard Fork",
             "Bitcoin forks into Bitcoin (BTC) and Bitcoin Cash (BCH). BCH gives holders "
             "free coins; short-term volatility ensues.",
             "Technology / Fork", "NEUTRAL", "Bitcoin Magazine / CoinDesk",
             "2700", "2700", "0", "", "hard fork,bitcoin cash,bch,scaling"),
    BTCEvent("2017-09-04", "China Bans ICOs and Exchanges",
             "Chinese government bans ICOs and orders all domestic cryptocurrency exchanges "
             "to shut down. BTC drops ~40%.",
             "Regulation", "BEARISH", "PBOC / Bloomberg",
             "4900", "3000", "-39", "", "china,ban,ico,exchange,regulation"),
    BTCEvent("2017-10-31", "CME Announces Bitcoin Futures",
             "CME Group announces it will launch Bitcoin futures, legitimizing Bitcoin as "
             "a tradable financial asset.",
             "Institutional / Derivatives", "BULLISH", "CME / CNBC",
             "6100", "6400", "+5", "", "futures,cme,institutional,derivatives"),
    BTCEvent("2017-12-17", "Bitcoin ATH — $19,783",
             "Bitcoin reaches its then-all-time high of $19,783. Media frenzy; retail "
             "FOMO at peak. Start of 2018 bear market.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "15000", "19783", "+32", "", "all time high,ath,bubble,retail"),
    BTCEvent("2017-12-18", "CME Bitcoin Futures Launch",
             "CME launches cash-settled Bitcoin futures. Institutional short selling begins; "
             "marks the top of the 2017 bull run.",
             "Institutional / Derivatives", "BEARISH", "CME",
             "19500", "18000", "-8", "", "futures,cme,top,derivatives"),

    # 2018
    BTCEvent("2018-01-16", "Global Crackdown on Cryptocurrency",
             "South Korea, China, US regulators all signal intent to regulate. Market-wide "
             "crash begins; BTC drops from $14K to under $10K.",
             "Regulation", "BEARISH", "Reuters / Bloomberg",
             "14000", "9500", "-32", "", "regulation,crackdown,korea,global"),
    BTCEvent("2018-01-26", "Coincheck Hack — $530M NEM Stolen",
             "Japanese exchange Coincheck is hacked for $530M in NEM tokens — the largest "
             "crypto theft ever at the time. Shakes market confidence in exchanges.",
             "Security / Hack", "BEARISH", "Coincheck / Reuters",
             "11000", "9700", "-12", "", "coincheck,hack,japan,nem,exchange,security"),
    BTCEvent("2018-03-07", "SEC Issues Subpoenas to ICO Projects",
             "SEC issues subpoenas to dozens of ICO projects, signaling aggressive "
             "enforcement and spooking the market.",
             "Regulation", "BEARISH", "SEC / WSJ",
             "10500", "8900", "-15", "", "sec,ico,regulation,enforcement"),
    BTCEvent("2018-03-14", "Google Bans Cryptocurrency Ads",
             "Google announces a ban on all cryptocurrency-related advertisements, "
             "following Facebook's similar ban in January 2018.",
             "Media / Regulation", "BEARISH", "Google / Bloomberg",
             "9200", "8200", "-11", "", "google,ads,ban,advertising,regulation"),
    BTCEvent("2018-06-10", "Coinrail Hack — Bitcoin Drops 10%",
             "South Korean exchange Coinrail is hacked for $40M in tokens. "
             "BTC drops 10% on news of another exchange breach.",
             "Security / Hack", "BEARISH", "Coinrail / Reuters",
             "7700", "6900", "-10", "", "coinrail,korea,hack,exchange,security"),
    BTCEvent("2018-09-20", "Goldman Sachs Abandons Bitcoin Trading Desk Plans",
             "Goldman Sachs reportedly shelves plans for a Bitcoin trading desk, "
             "disappointing institutional investors.",
             "Institutional", "BEARISH", "Goldman Sachs / Bloomberg",
             "6700", "6400", "-4", "", "goldman sachs,institutional,trading desk,investment bank"),
    BTCEvent("2018-11-15", "Bitcoin Cash Hash War",
             "Craig Wright's Bitcoin SV vs. Roger Ver's Bitcoin ABC hash war. Miners "
             "divert hash power; BTC drops 50% in weeks.",
             "Technology / Fork", "BEARISH", "CoinDesk",
             "6300", "3500", "-44", "", "hash war,bitcoin sv,bitcoin cash,fork"),
    BTCEvent("2018-12-15", "Bitcoin Year Low — $3,122",
             "Bitcoin bottoms at $3,122, an 84% decline from the $19,783 ATH. Crypto "
             "winter sets in.",
             "Market", "BEARISH", "CoinDesk",
             "3500", "3122", "-11", "", "bear market,crypto winter,bottom"),

    # 2019
    BTCEvent("2019-01-07", "Ethereum Constantinople Upgrade",
             "Ethereum completes its Constantinople upgrade, reducing block rewards. "
             "Broader crypto market begins 2019 recovery. BTC bottomed at $3,122 in Dec 2018.",
             "Competition / Technology", "NEUTRAL", "Ethereum Foundation",
             "3700", "3600", "-3", "", "ethereum,constantinople,upgrade,competition"),
    BTCEvent("2019-04-02", "Bitcoin Breaks $5,000 — Bull Signal",
             "Bitcoin suddenly surges from $4,100 to $5,000 in a single hour on April 2, "
             "signaling the end of the 2018 bear market.",
             "Market", "BULLISH", "CoinDesk",
             "4100", "5100", "+24", "", "breakout,5000,bull signal,recovery,2019"),
    BTCEvent("2019-05-14", "Bitcoin Surges Past $8,000",
             "Bitcoin continues its 2019 recovery rally, breaking $8,000 for the first "
             "time in over a year.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "7200", "8100", "+13", "", "milestone,$8000,rally,recovery"),
    BTCEvent("2019-06-18", "Facebook Announces Libra",
             "Facebook announces its Libra cryptocurrency project, bringing massive "
             "mainstream attention to crypto and Bitcoin.",
             "Institutional / Competition", "BULLISH", "Facebook / NYT",
             "8000", "11000", "+38", "", "facebook,libra,mainstream,adoption"),
    BTCEvent("2019-06-26", "Bitcoin Reaches $13,880 — 2019 High",
             "BTC rallies to $13,880, its highest since 2018 crash, driven by Libra "
             "announcement and renewed institutional interest.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "11000", "13880", "+26", "", "rally,2019 high,institutional"),
    BTCEvent("2019-07-11", "US Congress Holds Libra Hearing — Hostile",
             "US Congressional hearings on Facebook's Libra are highly hostile. "
             "Facebook faces massive pushback; Bitcoin drops from $13K.",
             "Regulation / Media", "BEARISH", "US Congress / Bloomberg",
             "13000", "10500", "-19", "", "libra,facebook,congress,regulation,hostile"),
    BTCEvent("2019-09-24", "Bitcoin Drops 20% in a Day",
             "BTC drops from $10,000 to $7,900 in a single day — one of the sharpest "
             "single-day drops of 2019, attributed to long liquidations.",
             "Market", "BEARISH", "CoinDesk",
             "10000", "7900", "-21", "", "crash,liquidation,2019,drop"),
    BTCEvent("2019-10-25", "Xi Jinping Endorses Blockchain Technology",
             "Chinese President Xi Jinping publicly endorses blockchain technology as "
             "an important breakthrough. Bitcoin surges 42% the next day.",
             "Political / Macro", "BULLISH", "Chinese State Media / CoinDesk",
             "7400", "10500", "+42", "", "china,xi jinping,blockchain,endorsement,political"),
    BTCEvent("2019-12-18", "Bitcoin Closes 2019 at $7,200",
             "Bitcoin ends 2019 at ~$7,200, up 87% from its 2018 low but down from the "
             "June high of $13,880.",
             "Market", "NEUTRAL", "CoinDesk",
             "7100", "7200", "+1", "", "year end,2019,recovery"),

    # 2020
    BTCEvent("2020-01-15", "Bitcoin Starts 2020 Strongly Above $9,000",
             "Bitcoin begins 2020 above $9,000, building momentum for the halving cycle. "
             "Institutional interest growing via OTC desks.",
             "Market", "BULLISH", "CoinDesk",
             "8700", "9100", "+5", "", "2020,year start,institutional,otc"),
    BTCEvent("2020-02-13", "Bitcoin Breaks $10,000",
             "Bitcoin crosses $10,000 before the halving, driven by Chinese stimulus "
             "measures amid early COVID-19 concerns.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "9500", "10400", "+9", "", "milestone,$10000,china,stimulus"),
    BTCEvent("2020-03-12", "COVID-19 Black Thursday Crash",
             "Global markets crash due to COVID-19 pandemic. Bitcoin drops from $9,000 "
             "to $3,800 in 24 hours (−57%), one of its worst single-day drops ever.",
             "Macro / Pandemic", "BEARISH", "CoinDesk / Bloomberg",
             "9000", "3800", "-57", "", "covid,pandemic,crash,black thursday,macro"),
    BTCEvent("2020-04-30", "Bitcoin Recovers to $9,000 Pre-Halving",
             "Bitcoin recovers from the COVID crash to $9,000 as the halving approaches. "
             "Hash rate also fully recovers.",
             "Market", "BULLISH", "CoinDesk",
             "5500", "9000", "+64", "", "recovery,halving,pre-halving,hash rate"),
    BTCEvent("2020-05-11", "Third Bitcoin Halving",
             "Block reward halves from 12.5 to 6.25 BTC. Triggers the 2020-2021 bull run "
             "that takes BTC from $8,500 to $69,000.",
             "Halving", "BULLISH", "Bitcoin Protocol",
             "8500", "8800", "+4", "", "halving,supply,mining,bull run"),
    BTCEvent("2020-07-27", "OCC Allows US Banks to Hold Crypto",
             "The Office of the Comptroller of the Currency issues guidance allowing "
             "national banks to provide cryptocurrency custody services.",
             "Regulation / Institutional", "BULLISH", "OCC / Bloomberg",
             "9700", "10500", "+8", "", "occ,banks,custody,regulation,institutional"),
    BTCEvent("2020-08-11", "MicroStrategy Buys $250M in Bitcoin",
             "MicroStrategy becomes the first public company to adopt Bitcoin as its "
             "primary treasury reserve asset. CEO Michael Saylor calls it digital gold.",
             "Institutional", "BULLISH", "MicroStrategy / Bloomberg",
             "11500", "12100", "+5", "", "microstrategy,institutional,corporate,treasury,michael saylor"),
    BTCEvent("2020-09-14", "MicroStrategy Buys Additional $175M Bitcoin",
             "MicroStrategy doubles down, purchasing an additional $175M in Bitcoin. "
             "Total holding exceeds $425M.",
             "Institutional", "BULLISH", "MicroStrategy / Bloomberg",
             "10700", "10900", "+2", "", "microstrategy,institutional,corporate,treasury"),
    BTCEvent("2020-10-08", "Square Buys $50M in Bitcoin",
             "Jack Dorsey's Square Inc. announces it purchased $50M in Bitcoin (~1% of "
             "total assets) as a treasury reserve.",
             "Institutional", "BULLISH", "Square / Bloomberg",
             "10700", "11300", "+6", "", "square,jack dorsey,institutional,corporate,treasury"),
    BTCEvent("2020-10-21", "PayPal Enables Bitcoin Buying",
             "PayPal announces all 346 million users can buy, sell, and hold Bitcoin "
             "directly. Massive mainstream adoption catalyst.",
             "Adoption", "BULLISH", "PayPal / CNBC",
             "12000", "13200", "+10", "", "paypal,mainstream,adoption,retail"),
    BTCEvent("2020-11-06", "Bitcoin Breaks $15,000 — 3-Year High",
             "Bitcoin crosses $15,000 for the first time since 2018, driven by PayPal, "
             "institutional buying, and dollar weakness.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "14000", "15500", "+11", "", "milestone,$15000,3 year high,institutional"),
    BTCEvent("2020-12-16", "Bitcoin Breaks 2017 ATH — $20,000",
             "Bitcoin breaks its 2017 all-time high of ~$20,000 for the first time, "
             "validating the new bull cycle.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "18000", "20000", "+11", "", "all time high,ath,milestone,breakout"),
    BTCEvent("2020-12-31", "Bitcoin Closes 2020 at $29,000",
             "Bitcoin ends 2020 at ~$29,000, a 300% annual gain. Best performing "
             "major asset class of the year.",
             "Market", "BULLISH", "CoinDesk",
             "27000", "29000", "+7", "", "year end,2020,annual return,best asset"),

    # 2021
    BTCEvent("2021-01-08", "Bitcoin Hits $40,000",
             "Bitcoin surges to $40,000 for the first time just one week after breaking "
             "$30,000. Fastest $10K increment ever.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "33000", "40000", "+21", "", "milestone,$40000,rally,speed"),
    BTCEvent("2021-01-29", "Elon Musk Adds #Bitcoin to Twitter Bio",
             "Elon Musk adds '#bitcoin' to his Twitter bio. BTC surges 20% in hours, "
             "demonstrating social media influence on price.",
             "Social Media / Influencer", "BULLISH", "Twitter / CoinDesk",
             "32000", "38000", "+19", "", "elon musk,twitter,social media,influencer"),
    BTCEvent("2021-02-08", "Tesla Buys $1.5B in Bitcoin",
             "Tesla announces it purchased $1.5B in Bitcoin and will accept it as payment. "
             "BTC surges to $44K.",
             "Institutional", "BULLISH", "Tesla / SEC Filing",
             "38000", "44000", "+16", "", "tesla,institutional,corporate,elon musk"),
    BTCEvent("2021-02-19", "Bitcoin Hits $52,640 — New ATH",
             "Bitcoin sets new all-time high driven by Tesla purchase, PayPal, "
             "institutional demand, and stimulus checks.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "47000", "52640", "+12", "", "all time high,ath,institutional,rally"),
    BTCEvent("2021-04-14", "Coinbase Direct Listing on NASDAQ",
             "Coinbase goes public on NASDAQ at $328/share (COIN). Bitcoin hits $64,000 "
             "same day — landmark for crypto legitimacy.",
             "Institutional / Exchange", "BULLISH", "NASDAQ / Bloomberg",
             "60000", "64895", "+8", "", "coinbase,ipo,nasdaq,institutional,adoption"),
    BTCEvent("2021-05-12", "Tesla Stops Accepting Bitcoin — Environmental",
             "Elon Musk announces Tesla will stop accepting Bitcoin due to environmental "
             "concerns over mining. BTC drops 15% immediately.",
             "Adoption / Environmental", "BEARISH", "Elon Musk / Twitter",
             "54000", "46000", "-15", "", "tesla,elon musk,environment,mining,esg"),
    BTCEvent("2021-05-19", "China Bans Bitcoin Mining",
             "China cracks down on Bitcoin mining and trading. Combined with Tesla news, "
             "BTC crashes from $64K to $30K (−53%).",
             "Regulation / Mining", "BEARISH", "PBOC / Bloomberg",
             "43000", "30000", "-30", "", "china,mining ban,regulation,crash"),
    BTCEvent("2021-07-21", "Bitcoin Bottoms at $29,700 then Recovers",
             "BTC hits cycle low of $29,700 then surges as hash rate recovers post-China "
             "ban and institutional buying resumes.",
             "Market", "BULLISH", "CoinDesk",
             "29700", "33000", "+11", "", "bottom,recovery,hash rate,institutional"),
    BTCEvent("2021-09-07", "El Salvador Makes Bitcoin Legal Tender",
             "El Salvador becomes the first country to adopt Bitcoin as legal tender. "
             "Chivo wallet launched. Short-term volatile but historically bullish.",
             "Regulation / Adoption", "BULLISH", "El Salvador Government / Reuters",
             "52000", "53000", "+2", "", "el salvador,legal tender,adoption,country"),
    BTCEvent("2021-10-20", "First Bitcoin Futures ETF (BITO) Launches",
             "ProShares Bitcoin Futures ETF (BITO) launches on NYSE — the first US-listed "
             "Bitcoin ETF. BTC hits new ATH of $67K.",
             "Institutional / ETF", "BULLISH", "ProShares / NYSE",
             "62000", "67000", "+8", "", "etf,bito,proshares,institutional,nyse"),
    BTCEvent("2021-11-10", "Bitcoin ATH — $69,000",
             "Bitcoin reaches its current all-time high of $68,789, driven by ETF launch, "
             "institutional demand, and inflation fears.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "67000", "68789", "+3", "", "all time high,ath,$69k,inflation"),
    BTCEvent("2021-12-04", "Bitcoin Flash Crash to $42,000",
             "Bitcoin flash crashes from $57K to $42K (-26%) in a cascade of leveraged "
             "liquidations totaling over $2B across all crypto.",
             "Market", "BEARISH", "CoinDesk",
             "57000", "42000", "-26", "", "flash crash,liquidation,leverage,cascade"),
    BTCEvent("2021-12-31", "Bitcoin Ends 2021 at $46,000",
             "Bitcoin ends 2021 at ~$46,000, up 60% for the year despite the crash from "
             "$69K ATH. Annual gain vs. S&P 500 (27%).",
             "Market", "BULLISH", "CoinDesk",
             "46000", "46211", "0", "", "year end,2021,annual return"),

    # 2022
    BTCEvent("2022-01-21", "Crypto Market Sell-Off on Fed Rate Hike Fears",
             "Federal Reserve signals aggressive rate hike cycle. Risk assets sell off; "
             "BTC drops from $47K to $33K.",
             "Macro / Fed", "BEARISH", "Fed / Bloomberg",
             "47000", "33000", "-30", "", "fed,rate hike,macro,inflation,risk off"),
    BTCEvent("2022-03-16", "Fed Hikes Rates for First Time Since 2018",
             "Federal Reserve raises rates by 25bps — the first hike since 2018. "
             "Risk assets initially rally on 'buy the news' before resuming decline.",
             "Macro / Fed", "BEARISH", "Federal Reserve / Bloomberg",
             "38000", "41000", "+8", "", "fed,rate hike,monetary policy,macro"),
    BTCEvent("2022-04-11", "Bitcoin Drops Below $40,000",
             "Bitcoin falls below $40,000 as Fed tightening accelerates. Correlation "
             "with NASDAQ at historic highs.",
             "Market", "BEARISH", "CoinDesk",
             "43000", "39500", "-8", "", "bear,40000,nasdaq,correlation,fed"),
    BTCEvent("2022-05-09", "Luna/UST Collapse",
             "Terra Luna ecosystem collapses; UST de-pegs from $1 and Luna goes to zero. "
             "$60B wiped out; BTC drops from $36K to $26K.",
             "DeFi / Stablecoin", "BEARISH", "CoinDesk / Bloomberg",
             "36000", "26000", "-28", "", "luna,terra,ust,stablecoin,depeg,collapse"),
    BTCEvent("2022-06-13", "Celsius Pauses Withdrawals",
             "Celsius Network halts all withdrawals and transfers citing extreme market "
             "conditions. BTC drops below $20K.",
             "Exchange / Lending", "BEARISH", "Celsius / Bloomberg",
             "28000", "20000", "-29", "", "celsius,withdrawal freeze,lending,contagion"),
    BTCEvent("2022-06-18", "Three Arrows Capital Collapse",
             "Hedge fund Three Arrows Capital (3AC) defaults on loans worth $670M. "
             "Crypto credit crisis deepens.",
             "Institutional / Contagion", "BEARISH", "WSJ / Bloomberg",
             "20000", "17500", "-13", "", "three arrows,3ac,hedge fund,contagion,liquidation"),
    BTCEvent("2022-07-13", "Bitcoin Bottom — $17,593",
             "Bitcoin hits $17,593, the lowest since December 2020. Miners capitulate; "
             "hash ribbons signal extreme fear.",
             "Market", "BEARISH", "CoinDesk",
             "19000", "17593", "-7", "", "bottom,bear market,miners,capitulation"),
    BTCEvent("2022-11-08", "FTX Liquidity Crisis",
             "Binance CEO CZ tweets concern over FTT tokens; FTX faces bank run. Binance "
             "backs out of rescue deal. BTC drops from $21K to $16K.",
             "Exchange / Fraud", "BEARISH", "CoinDesk / WSJ",
             "21000", "16000", "-24", "", "ftx,sam bankman-fried,sbf,exchange,fraud,contagion"),
    BTCEvent("2022-11-11", "FTX Files Bankruptcy — Sam Bankman-Fried Arrested",
             "FTX files for Chapter 11 bankruptcy. SBF arrested. $8B in customer funds "
             "missing. Largest crypto exchange collapse. BTC hits $15,600.",
             "Exchange / Fraud", "BEARISH", "FTX / DOJ",
             "16000", "15600", "-3", "", "ftx,sbf,bankruptcy,fraud,contagion"),

    # 2023
    BTCEvent("2023-01-12", "Bitcoin Starts 2023 Recovery",
             "Bitcoin rallies from $16K lows, starting the 2023 recovery rally amid "
             "expectations of Fed pivot and post-FTX reset.",
             "Market", "BULLISH", "CoinDesk",
             "17000", "21000", "+24", "", "recovery,2023,bear market bottom,fed pivot"),
    BTCEvent("2023-02-16", "Kraken Shuts US Staking Program — SEC Settlement",
             "Kraken settles with SEC for $30M and shuts its US crypto staking service. "
             "Signals SEC targeting crypto yield products.",
             "Regulation", "BEARISH", "SEC / Kraken",
             "21500", "21000", "-2", "", "kraken,staking,sec,regulation,enforcement"),
    BTCEvent("2023-03-10", "Silicon Valley Bank Collapse",
             "SVB collapses in the second-largest US bank failure ever. Bitcoin initially "
             "drops then surges as banking fears resurface. USDC briefly de-pegs.",
             "Macro / Banking", "BULLISH", "FDIC / Bloomberg",
             "20000", "22000", "+10", "", "svb,banking crisis,safe haven,macro,usdc"),
    BTCEvent("2023-03-22", "Fed Hikes 25bps Amid Banking Crisis",
             "Fed hikes despite banking sector stress, signaling commitment to fighting "
             "inflation. Bitcoin holds above $28K showing resilience.",
             "Macro / Fed", "NEUTRAL", "Federal Reserve / Bloomberg",
             "27500", "28000", "+2", "", "fed,rate hike,banking crisis,macro,resilience"),
    BTCEvent("2023-04-13", "Bitcoin Crosses $30,000",
             "Bitcoin breaks $30,000 for the first time since June 2022 as banking "
             "fears boost safe haven demand.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "28000", "30500", "+9", "", "milestone,$30000,safe haven,recovery"),
    BTCEvent("2023-06-05", "SEC Sues Binance and Coinbase",
             "SEC files suits against Binance (13 charges) and Coinbase (1 day later) "
             "for operating unregistered exchanges. BTC drops 5%.",
             "Regulation", "BEARISH", "SEC / DOJ",
             "27000", "25700", "-5", "", "sec,binance,coinbase,lawsuit,regulation,enforcement"),
    BTCEvent("2023-06-15", "BlackRock Files for Bitcoin Spot ETF",
             "BlackRock, the world's largest asset manager ($10T AUM), files for a spot "
             "Bitcoin ETF. Considered the most credible application to date.",
             "Institutional / ETF", "BULLISH", "BlackRock / SEC",
             "25000", "30000", "+20", "", "blackrock,etf,spot etf,institutional"),
    BTCEvent("2023-08-17", "Bitcoin Drops to $26,000 on Rising Yields",
             "US 10-year Treasury yields hit 4.3% (15-year high), pressuring risk assets. "
             "Bitcoin drops from $29K to $26K.",
             "Macro", "BEARISH", "Bloomberg",
             "29000", "26000", "-10", "", "yields,treasury,macro,risk off,rates"),
    BTCEvent("2023-10-17", "SEC Delays Bitcoin ETF Decisions Again",
             "SEC delays decisions on multiple spot Bitcoin ETF applications, "
             "temporarily disappointing markets.",
             "Regulation / ETF", "BEARISH", "SEC / CoinDesk",
             "28000", "26900", "-4", "", "etf,sec,delay,regulation"),
    BTCEvent("2023-10-23", "Bitcoin Surges on ETF Approval Rumors",
             "Bitcoin surges to $35K on fake news of BlackRock ETF approval, then "
             "partially retraces but holds gains on real anticipation.",
             "ETF / Media", "BULLISH", "CoinDesk / Bloomberg",
             "27000", "35000", "+30", "", "etf,blackrock,approval,rumor"),
    BTCEvent("2023-12-05", "Bitcoin Hits $44,000 — 19-Month High",
             "Bitcoin breaks $44,000 for the first time since April 2022, driven by "
             "ETF anticipation and favorable macro conditions.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "37000", "44000", "+19", "", "milestone,$44000,19 month high,etf,rally"),

    # 2024
    BTCEvent("2024-01-03", "Bitcoin's 15th Anniversary — Network Celebrates",
             "Bitcoin celebrates its 15th birthday. Network hash rate at all-time highs; "
             "BTC holds above $44K with ETF decision imminent.",
             "Technology", "BULLISH", "Bitcoin Magazine",
             "44000", "44500", "+1", "", "anniversary,15 years,hash rate,network"),
    BTCEvent("2024-01-10", "SEC Approves 11 Spot Bitcoin ETFs",
             "SEC simultaneously approves 11 spot Bitcoin ETFs including BlackRock IBIT, "
             "Fidelity FBTC. First day sees $4.6B in trading volume.",
             "Institutional / ETF", "BULLISH", "SEC / Bloomberg",
             "46000", "47000", "+2", "", "spot etf,sec,blackrock,fidelity,ibit,institutional"),
    BTCEvent("2024-02-14", "Bitcoin ETF Inflows Surge — IBIT Record",
             "BlackRock's IBIT ETF records $612M in a single day, the largest daily "
             "inflow of any ETF ever at the time.",
             "Institutional / ETF", "BULLISH", "BlackRock / Bloomberg",
             "52000", "54000", "+4", "", "etf,blackrock,ibit,inflows,institutional"),
    BTCEvent("2024-03-05", "Bitcoin Breaks 2021 ATH — $69,000+",
             "Bitcoin surpasses its 2021 all-time high of $69K for the first time, "
             "driven by ETF inflows and halving anticipation.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "62000", "69200", "+12", "", "all time high,ath,breakout,etf,halving"),
    BTCEvent("2024-04-19", "Fourth Bitcoin Halving",
             "Block reward halves from 6.25 to 3.125 BTC at block 840,000. Miner fees "
             "spike to record highs due to Runes protocol launch.",
             "Halving", "BULLISH", "Bitcoin Protocol",
             "64000", "64000", "0", "", "halving,supply,mining,runes,ordinals"),
    BTCEvent("2024-05-23", "SEC Approves Spot Ethereum ETFs",
             "SEC unexpectedly approves spot Ethereum ETFs, signaling regulatory shift "
             "and benefiting the broader crypto market including BTC.",
             "Institutional / ETF / Regulation", "BULLISH", "SEC / Bloomberg",
             "67000", "69000", "+3", "", "ethereum,etf,sec,regulation,approval"),
    BTCEvent("2024-10-14", "Bitcoin Surges to $68K — Election Optimism",
             "Bitcoin rises to $68K driven by expectations that a Trump presidency would "
             "be pro-crypto.",
             "Political / Macro", "BULLISH", "CoinDesk",
             "62000", "68000", "+10", "", "trump,election,political,regulation,pro-crypto"),
    BTCEvent("2024-11-06", "Trump Wins US Presidential Election",
             "Donald Trump wins the 2024 US election. Bitcoin surges past $75K as markets "
             "price in crypto-friendly policies and potential strategic reserve.",
             "Political / Regulation", "BULLISH", "AP News / CoinDesk",
             "68000", "75000", "+10", "", "trump,election,president,strategic reserve,pro-crypto"),
    BTCEvent("2024-11-22", "Bitcoin Breaks $100,000 for First Time",
             "Bitcoin surpasses $100,000 for the first time, driven by Trump victory, "
             "ETF inflows, and halving supply squeeze.",
             "Price Milestone", "BULLISH", "CoinDesk / Bloomberg",
             "90000", "103000", "+14", "", "milestone,$100k,all time high,ath"),
    BTCEvent("2024-12-05", "Bitcoin Hits $104,000 — New ATH",
             "Bitcoin sets a new all-time high of $104,000 as total ETF AUM crosses "
             "$100B and institutional adoption accelerates.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "100000", "104000", "+4", "", "all time high,ath,$104k,etf,institutional"),
    BTCEvent("2024-12-19", "Fed Rate Cut — Hawkish Guidance Sends BTC to $93K",
             "Fed cuts rates 25bps but signals only 2 cuts in 2025 (vs. 4 expected). "
             "Risk assets sell off; BTC drops from $107K to $93K.",
             "Macro / Fed", "BEARISH", "Federal Reserve / Bloomberg",
             "107000", "93000", "-13", "", "fed,rate cut,hawkish,macro,guidance"),

    # 2025
    BTCEvent("2025-01-20", "Trump Signs Crypto Executive Order",
             "President Trump signs an executive order on Day 1 directing the creation of "
             "a Presidential Digital Asset Working Group and potential strategic BTC reserve.",
             "Regulation / Political", "BULLISH", "White House / CoinDesk",
             "102000", "107000", "+5", "", "trump,executive order,strategic reserve,regulation"),
    BTCEvent("2025-01-23", "Bitcoin Hits New ATH Above $109,000",
             "Bitcoin sets a new all-time high above $109,000 on Trump inauguration day, "
             "driven by executive order expectations and ETF inflows.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "105000", "109000", "+4", "", "all time high,ath,$109k,trump,inauguration"),
    BTCEvent("2025-02-03", "Bitcoin Drops to $91,000 on Tariff Fears",
             "Trump announces tariffs on Canada, Mexico, and China. Risk assets sell off; "
             "BTC drops from $105K to $91K in 48 hours.",
             "Macro / Political", "BEARISH", "Bloomberg / CoinDesk",
             "104000", "91000", "-13", "", "tariffs,trade war,macro,trump,political"),
    BTCEvent("2025-03-07", "US Strategic Bitcoin Reserve Announced",
             "Trump signs executive order establishing a US Strategic Bitcoin Reserve "
             "using seized government BTC (~200,000 BTC). BTC surges on the news.",
             "Regulation / Government", "BULLISH", "White House / Bloomberg",
             "88000", "95000", "+8", "", "strategic reserve,government,us,trump,adoption"),
    BTCEvent("2025-03-11", "Bitcoin Drops Below $80,000 — Macro Pressure",
             "Bitcoin falls below $80,000 as recession fears and global trade war "
             "escalation weigh on risk assets. Strategic reserve euphoria fades.",
             "Macro / Market", "BEARISH", "Bloomberg / CoinDesk",
             "90000", "78000", "-13", "", "recession,trade war,macro,risk off,correction"),
]


# ---------------------------------------------------------------------------
# LIVE SCRAPERS (supplemental — adds recent news)
# ---------------------------------------------------------------------------

def scrape_coindesk_news(max_articles: int = 20) -> list[BTCEvent]:
    """Scrape recent Bitcoin headlines from CoinDesk."""
    events: list[BTCEvent] = []
    url = "https://www.coindesk.com/tag/bitcoin"
    try:
        log.info("Scraping CoinDesk...")
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "lxml")

        articles = soup.select("a[data-testid='card-title-link'], "
                               "h4 a, h3 a, .article-card a, "
                               "[class*='headline'] a")[:max_articles]
        seen: set[str] = set()
        for a in articles:
            title = a.get_text(strip=True)
            href = a.get("href", "")
            if not title or title in seen or len(title) < 20:
                continue
            seen.add(title)
            full_url = href if href.startswith("http") else f"https://www.coindesk.com{href}"
            impact = classify_impact(title)
            events.append(BTCEvent(
                date=datetime.today().strftime("%Y-%m-%d"),
                title=title[:200],
                description="[Scraped from CoinDesk — description requires full article fetch]",
                category=infer_category(title),
                impact=impact,
                source="CoinDesk",
                url=full_url,
                tags=extract_tags(title),
            ))
        log.info(f"CoinDesk: {len(events)} articles found")
    except Exception as e:
        log.warning(f"CoinDesk scrape failed: {e}")
    return events


def scrape_bitcoin_magazine(max_articles: int = 20) -> list[BTCEvent]:
    """Scrape recent news from Bitcoin Magazine."""
    events: list[BTCEvent] = []
    url = "https://bitcoinmagazine.com/articles"
    try:
        log.info("Scraping Bitcoin Magazine...")
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "lxml")

        articles = soup.select("h2 a, h3 a, .article__title a, "
                               "[class*='title'] a")[:max_articles]
        seen: set[str] = set()
        for a in articles:
            title = a.get_text(strip=True)
            href = a.get("href", "")
            if not title or title in seen or len(title) < 20:
                continue
            seen.add(title)
            full_url = href if href.startswith("http") else f"https://bitcoinmagazine.com{href}"
            events.append(BTCEvent(
                date=datetime.today().strftime("%Y-%m-%d"),
                title=title[:200],
                description="[Scraped from Bitcoin Magazine]",
                category=infer_category(title),
                impact=classify_impact(title),
                source="Bitcoin Magazine",
                url=full_url,
                tags=extract_tags(title),
            ))
        log.info(f"Bitcoin Magazine: {len(events)} articles found")
    except Exception as e:
        log.warning(f"Bitcoin Magazine scrape failed: {e}")
    return events


def scrape_cointelegraph(max_articles: int = 20) -> list[BTCEvent]:
    """Scrape recent Bitcoin news from CoinTelegraph."""
    events: list[BTCEvent] = []
    url = "https://cointelegraph.com/tags/bitcoin"
    try:
        log.info("Scraping CoinTelegraph...")
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "lxml")

        articles = soup.select("a.post-card__title-link, h2 a, h3 a, "
                               "[class*='post-card'] a")[:max_articles]
        seen: set[str] = set()
        for a in articles:
            title = a.get_text(strip=True)
            href = a.get("href", "")
            if not title or title in seen or len(title) < 20:
                continue
            seen.add(title)
            full_url = href if href.startswith("http") else f"https://cointelegraph.com{href}"
            events.append(BTCEvent(
                date=datetime.today().strftime("%Y-%m-%d"),
                title=title[:200],
                description="[Scraped from CoinTelegraph]",
                category=infer_category(title),
                impact=classify_impact(title),
                source="CoinTelegraph",
                url=full_url,
                tags=extract_tags(title),
            ))
        log.info(f"CoinTelegraph: {len(events)} articles found")
    except Exception as e:
        log.warning(f"CoinTelegraph scrape failed: {e}")
    return events


# ---------------------------------------------------------------------------
# NLP HELPERS
# ---------------------------------------------------------------------------

BULLISH_KEYWORDS = {
    "buy", "bull", "surge", "rally", "soar", "jump", "rise", "gain", "record",
    "ath", "all-time high", "approve", "approval", "adopt", "adoption", "launch",
    "invest", "institutional", "etf", "halving", "legal", "green", "recover",
    "breakout", "milestone", "fund", "reserve", "trump", "pro-crypto", "listing",
    "upgrade", "partnership", "integration", "accept",
}

BEARISH_KEYWORDS = {
    "sell", "bear", "crash", "drop", "fall", "plunge", "dump", "ban", "hack",
    "stolen", "theft", "fraud", "scam", "bankrupt", "suspend", "halt", "block",
    "crackdown", "restrict", "fine", "arrest", "collapse", "lose", "loss",
    "fear", "panic", "FUD", "warning", "risk", "regulation", "sue", "lawsuit",
    "reject", "denied", "debt", "contagion", "liquidat",
}


def classify_impact(text: str) -> str:
    text_lower = text.lower()
    bull = sum(1 for k in BULLISH_KEYWORDS if k in text_lower)
    bear = sum(1 for k in BEARISH_KEYWORDS if k in text_lower)
    if bull > bear:
        return "BULLISH"
    if bear > bull:
        return "BEARISH"
    return "NEUTRAL"


CATEGORY_MAP = {
    "hack|theft|stolen|breach|exploit": "Security / Hack",
    "etf|fund|futures|derivative|option": "Institutional / ETF",
    "halving|block reward|mining|hash": "Halving / Mining",
    "regulation|ban|law|legal|sec|cftc|fca|pboc|government": "Regulation",
    "exchange|coinbase|binance|kraken|ftx|bybit|okx": "Exchange",
    "macro|fed|inflation|rate|recession|bank|crisis": "Macro",
    "fork|upgrade|protocol|segwit|taproot|lightning": "Technology / Fork",
    "institutional|corporate|microstrategy|tesla|blackrock": "Institutional",
    "adoption|payment|accept|merchant": "Adoption",
    "defi|stablecoin|usdc|usdt|dai|ust": "DeFi / Stablecoin",
    "social|twitter|tweet|elon|influencer": "Social Media / Influencer",
    "political|president|election|congress|senate": "Political",
}


def infer_category(text: str) -> str:
    text_lower = text.lower()
    for pattern, category in CATEGORY_MAP.items():
        if re.search(pattern, text_lower):
            return category
    return "Market"


TAG_KEYWORDS = [
    "bitcoin", "btc", "etf", "sec", "halving", "institutional", "regulation",
    "exchange", "mining", "defi", "stablecoin", "macro", "fed", "trump",
    "blackrock", "coinbase", "tesla", "china", "fork", "lightning", "taproot",
    "crash", "rally", "adoption", "ban", "hack", "fraud",
]


def extract_tags(text: str) -> str:
    text_lower = text.lower()
    return ",".join(t for t in TAG_KEYWORDS if t in text_lower)


# ---------------------------------------------------------------------------
# EXCEL BUILDER
# ---------------------------------------------------------------------------

CATEGORY_COLORS = {
    "Halving": "FFF3CD",
    "Halving / Mining": "FFF3CD",
    "Price Milestone": "D4EDDA",
    "Institutional": "CCE5FF",
    "Institutional / ETF": "CCE5FF",
    "Institutional / Derivatives": "CCE5FF",
    "Institutional / Exchange": "CCE5FF",
    "Institutional / Contagion": "F8D7DA",
    "Regulation": "E2E3E5",
    "Regulation / Adoption": "E2E3E5",
    "Regulation / Law": "E2E3E5",
    "Regulation / ETF": "E2E3E5",
    "Regulation / Mining": "E2E3E5",
    "Regulation / Government": "E2E3E5",
    "Regulation / Political": "E2E3E5",
    "Security / Hack": "F8D7DA",
    "Exchange / Hack": "F8D7DA",
    "Exchange / Lending": "F8D7DA",
    "Exchange / Fraud": "F8D7DA",
    "Technology / Fork": "E8D5F5",
    "Macro": "FFE5D0",
    "Macro / Banking": "FFE5D0",
    "Macro / Pandemic": "FFE5D0",
    "Macro / Fed": "FFE5D0",
    "DeFi / Stablecoin": "FCE4EC",
    "Social Media / Influencer": "E0F7FA",
    "Adoption": "F0FFF0",
    "Market": "F5F5F5",
    "Political": "FFF8E1",
    "Political / Regulation": "FFF8E1",
    "Political / Macro": "FFF8E1",
    "Competition": "FAFAFA",
    "Media": "E3F2FD",
}

IMPACT_COLORS = {
    "BULLISH": "C6EFCE",
    "BEARISH": "FFC7CE",
    "NEUTRAL": "FFEB9C",
}

THIN = Side(border_style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("Date", 14),
    ("Title", 55),
    ("Category", 28),
    ("Impact", 10),
    ("Price Before ($)", 17),
    ("Price After ($)", 16),
    ("Change (%)", 12),
    ("Description", 80),
    ("Source", 25),
    ("URL", 50),
    ("Tags", 40),
]


def _cell_fill(color_hex: str) -> PatternFill:
    return PatternFill("solid", fgColor=color_hex)


def build_excel(events: list[BTCEvent], output_path: str) -> None:
    log.info(f"Building Excel file: {output_path}")
    df = _events_to_df(events)

    # ---- Main events sheet ----
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="BTC Events", index=False)
        _style_events_sheet(writer.book["BTC Events"], df)

        # ---- Statistics sheet ----
        stats_df = _build_stats(df)
        stats_df.to_excel(writer, sheet_name="Statistics", index=False)
        _style_stats_sheet(writer.book["Statistics"])

        # ---- Category summary ----
        cat_df = _build_category_summary(df)
        cat_df.to_excel(writer, sheet_name="Category Summary", index=False)

        # ---- Prediction features ----
        feat_df = _build_prediction_features(df)
        feat_df.to_excel(writer, sheet_name="Prediction Features", index=False)
        _style_prediction_sheet(writer.book["Prediction Features"])

    log.info(f"Excel saved: {output_path}  ({len(events)} events)")


def _events_to_df(events: list[BTCEvent]) -> pd.DataFrame:
    rows = []
    for e in events:
        rows.append({
            "Date": e.date,
            "Title": e.title,
            "Category": e.category,
            "Impact": e.impact,
            "Price Before ($)": e.price_before,
            "Price After ($)": e.price_after,
            "Change (%)": e.price_change_pct,
            "Description": e.description,
            "Source": e.source,
            "URL": e.url,
            "Tags": e.tags,
        })
    df = pd.DataFrame(rows)
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.strftime("%Y-%m-%d")
    df = df.sort_values("Date").reset_index(drop=True)
    return df


def _style_events_sheet(ws, df: pd.DataFrame) -> None:
    # Header row
    header_fill = _cell_fill("1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Data rows
    col_names = [c[0] for c in COLUMNS]
    for row_idx, row in df.iterrows():
        excel_row = row_idx + 2
        category = str(row.get("Category", ""))
        impact = str(row.get("Impact", ""))
        cat_color = CATEGORY_COLORS.get(category, "FFFFFF")
        impact_color = IMPACT_COLORS.get(impact, "FFFFFF")

        for col_idx, col_name in enumerate(col_names, start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.value = row.get(col_name, "")
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(col_name in ("Description", "Title", "Tags")),
            )
            if col_name == "Impact":
                cell.fill = _cell_fill(impact_color)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_name == "Category":
                cell.fill = _cell_fill(cat_color)
            elif col_name == "Change (%)":
                val = str(cell.value).replace("%", "").strip()
                try:
                    num = float(val)
                    cell.fill = _cell_fill("C6EFCE" if num >= 0 else "FFC7CE")
                    cell.font = Font(bold=True)
                except ValueError:
                    pass
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_name == "URL" and str(cell.value).startswith("http"):
                cell.hyperlink = str(cell.value)
                cell.font = Font(color="0563C1", underline="single")
            else:
                cell.fill = _cell_fill(cat_color)

        ws.row_dimensions[excel_row].height = 40

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions


def _build_stats(df: pd.DataFrame) -> pd.DataFrame:
    numeric_change = pd.to_numeric(
        df["Change (%)"].astype(str).str.replace("%", "").str.replace("+", ""),
        errors="coerce"
    )
    df = df.copy()
    df["_change"] = numeric_change

    stats = {
        "Metric": [
            "Total Events",
            "Bullish Events",
            "Bearish Events",
            "Neutral Events",
            "Avg Bullish Move (%)",
            "Avg Bearish Move (%)",
            "Largest Single Rally (%)",
            "Largest Single Drop (%)",
            "Most Common Category",
            "Events per Year (avg)",
            "Date Range Start",
            "Date Range End",
        ],
        "Value": [
            len(df),
            len(df[df["Impact"] == "BULLISH"]),
            len(df[df["Impact"] == "BEARISH"]),
            len(df[df["Impact"] == "NEUTRAL"]),
            round(df[df["_change"] > 0]["_change"].mean(), 1) if not df[df["_change"] > 0].empty else 0,
            round(df[df["_change"] < 0]["_change"].mean(), 1) if not df[df["_change"] < 0].empty else 0,
            round(df["_change"].max(), 1),
            round(df["_change"].min(), 1),
            df["Category"].value_counts().idxmax() if not df.empty else "N/A",
            round(len(df) / max(1, len(df["Date"].str[:4].unique())), 1),
            df["Date"].min(),
            df["Date"].max(),
        ],
    }
    return pd.DataFrame(stats)


def _style_stats_sheet(ws) -> None:
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    header_fill = _cell_fill("1F3864")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER


def _build_category_summary(df: pd.DataFrame) -> pd.DataFrame:
    summary = df.groupby(["Category", "Impact"]).size().reset_index(name="Count")
    return summary.sort_values("Count", ascending=False)


def _build_prediction_features(df: pd.DataFrame) -> pd.DataFrame:
    """
    Outputs a machine-learning ready feature table.
    Each row = one event with numeric features useful for prediction models.
    """
    df = df.copy()
    df["change_num"] = pd.to_numeric(
        df["Change (%)"].astype(str).str.replace("%", "").str.replace("+", ""),
        errors="coerce"
    )
    df["is_halving"] = df["Category"].str.contains("Halving", na=False).astype(int)
    df["is_regulation"] = df["Category"].str.contains("Regulation", na=False).astype(int)
    df["is_hack"] = df["Category"].str.contains("Hack|Security|Fraud", na=False).astype(int)
    df["is_institutional"] = df["Category"].str.contains("Institutional", na=False).astype(int)
    df["is_macro"] = df["Category"].str.contains("Macro", na=False).astype(int)
    df["is_exchange"] = df["Category"].str.contains("Exchange", na=False).astype(int)
    df["is_etf"] = df["Tags"].str.contains("etf", na=False).astype(int)
    df["is_china"] = df["Tags"].str.contains("china", na=False).astype(int)
    df["is_us_gov"] = df["Tags"].str.contains("trump|sec|regulation|government", na=False).astype(int)
    df["impact_numeric"] = df["Impact"].map({"BULLISH": 1, "NEUTRAL": 0, "BEARISH": -1})

    features = df[[
        "Date", "Title", "Category", "Impact", "impact_numeric",
        "change_num",
        "is_halving", "is_regulation", "is_hack", "is_institutional",
        "is_macro", "is_exchange", "is_etf", "is_china", "is_us_gov",
    ]].rename(columns={"change_num": "price_change_pct"})
    return features


def _style_prediction_sheet(ws) -> None:
    header_fill = _cell_fill("1F3864")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 28
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main() -> None:
    log.info("=== Bitcoin Price Events Scraper ===")

    all_events: list[BTCEvent] = list(HISTORICAL_EVENTS)
    log.info(f"Loaded {len(all_events)} curated historical events")

    # Scrape live news sources
    for scraper in [scrape_coindesk_news, scrape_bitcoin_magazine, scrape_cointelegraph]:
        live_events = scraper()
        all_events.extend(live_events)
        time.sleep(2)  # polite delay between requests

    # Deduplicate by title
    seen_titles: set[str] = set()
    unique_events: list[BTCEvent] = []
    for e in all_events:
        key = e.title.lower().strip()[:80]
        if key not in seen_titles:
            seen_titles.add(key)
            unique_events.append(e)

    log.info(f"Total unique events after deduplication: {len(unique_events)}")

    output_path = f"/Users/noledge/Insider-Trading/Insider_Trading/{OUTPUT_FILE}"
    build_excel(unique_events, output_path)

    # Print summary
    bullish = sum(1 for e in unique_events if e.impact == "BULLISH")
    bearish = sum(1 for e in unique_events if e.impact == "BEARISH")
    neutral = sum(1 for e in unique_events if e.impact == "NEUTRAL")

    print(f"\n{'='*55}")
    print(f"  Bitcoin Events Database — Build Complete")
    print(f"{'='*55}")
    print(f"  Total events  : {len(unique_events)}")
    print(f"  Bullish       : {bullish}")
    print(f"  Bearish       : {bearish}")
    print(f"  Neutral       : {neutral}")
    print(f"  Output file   : {output_path}")
    print(f"  Sheets        : BTC Events | Statistics | Category Summary | Prediction Features")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
